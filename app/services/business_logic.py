import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Callable

import openpyxl
from app.config import RESOURCE_DIR
from deep_translator import GoogleTranslator
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


DATE_SUFFIX_PATTERN = re.compile(r"\s+(\d{2})[./-]?(\d{2})[./-]?(\d{4})$")
DOCUMENT_CODE_PATTERN = re.compile(r"^(?P<code>\d+[^\W_]*(?:[-/.][^\W_]+)+)(?:\s+(?P<content>.*))?$")
SHEET_NAME = "Quản lý văn bản"
TRANSLATE_RULE_FILE = "translate-rule.xlsx"
HEADERS = [
    "Số",
    "Ngày ký",
    "Trích yếu nội dung",
    "Trích yếu tiếng Anh",
    "Người xin số",
    "CBTT",
    "BBH\nLiên quan",
]

TRICH_YEU_COL = 3
TRICH_YEU_EN_COL = 4
TOTAL_COLUMNS = len(HEADERS)

ProgressCallback = Callable[[str], None]


@dataclass
class ProcessingResult:
    ok: bool
    message: str
    output_path: str = ""
    total_files: int = 0
    valid_files: int = 0
    invalid_files: int = 0
    rows_added: int = 0
    duplicate_rows: int = 0
    invalid_items: list[dict] = field(default_factory=list)
    system_error: str = ""


def _notify(progress_callback: ProgressCallback | None, message: str) -> None:
    if progress_callback:
        progress_callback(message)


def chuan_hoa_gia_tri(value) -> str:
    return str(value or "").strip()


def chuan_hoa_header(value) -> str:
    return re.sub(r"\s+", " ", chuan_hoa_gia_tri(value).lower()).strip()


def tach_ngay_ky_tu_ten_file(ten_file_khong_duoi: str) -> tuple[str, str]:
    match = DATE_SUFFIX_PATTERN.search(ten_file_khong_duoi)
    if not match:
        return ten_file_khong_duoi.strip(), ""

    ngay, thang, nam = match.groups()
    try:
        datetime.strptime(f"{ngay}{thang}{nam}", "%d%m%Y")
    except ValueError:
        return ten_file_khong_duoi.strip(), ""

    ten_file_khong_ngay = ten_file_khong_duoi[: match.start()].strip(" _-.")
    return ten_file_khong_ngay, f"{ngay}/{thang}/{nam}"


def tim_file_quy_tac_dich() -> Path:
    candidates = [Path.cwd() / TRANSLATE_RULE_FILE]

    if getattr(sys, "frozen", False):
        candidates.append(RESOURCE_DIR / TRANSLATE_RULE_FILE)
        candidates.append(Path(sys.executable).resolve().parent / TRANSLATE_RULE_FILE)
    else:
        candidates.append(Path(__file__).resolve().parents[2] / TRANSLATE_RULE_FILE)

    da_kiem_tra = set()
    for path in candidates:
        resolved = path.resolve()
        if resolved in da_kiem_tra:
            continue
        da_kiem_tra.add(resolved)
        if resolved.exists():
            return resolved

    return candidates[0].resolve()


def la_header_quy_tac_dich(tieng_viet, tieng_anh) -> bool:
    tieng_viet = chuan_hoa_header(tieng_viet)
    tieng_anh = chuan_hoa_header(tieng_anh)
    return tieng_viet in {"tiếng việt", "tieng viet", "vietnamese", "vi"} or tieng_anh in {
        "tiếng anh",
        "tieng anh",
        "english",
        "en",
    }


def chuan_hoa_khoa_quy_tac_dich(value) -> str:
    return re.sub(r"\s+", " ", chuan_hoa_gia_tri(value).lower()).strip()


def tao_mau_quy_tac_dich(tieng_viet: str) -> str:
    parts = [re.escape(part) for part in re.split(r"\s+", tieng_viet.strip()) if part]
    return r"\s+".join(parts)


def la_ky_tu_trong_tu(value: str) -> bool:
    return bool(value and re.match(r"[^\W_]", value, re.UNICODE))


@lru_cache(maxsize=1)
def doc_quy_tac_dich() -> tuple:
    rule_path = tim_file_quy_tac_dich()
    if not rule_path.exists():
        return ()

    try:
        wb = openpyxl.load_workbook(rule_path, data_only=True, read_only=True)
    except Exception:
        return ()

    rules_by_key = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            tieng_viet = chuan_hoa_gia_tri(row[0] if len(row) > 0 else "")
            tieng_anh = chuan_hoa_gia_tri(row[1] if len(row) > 1 else "")

            if not tieng_viet or not tieng_anh or la_header_quy_tac_dich(tieng_viet, tieng_anh):
                continue

            rules_by_key[chuan_hoa_khoa_quy_tac_dich(tieng_viet)] = (
                tieng_viet,
                tieng_anh,
                re.compile(tao_mau_quy_tac_dich(tieng_viet), re.IGNORECASE),
            )

    return tuple(sorted(rules_by_key.values(), key=lambda item: len(item[0]), reverse=True))


def dinh_dang_so_van_ban_de_ghi(value) -> str:
    value = chuan_hoa_gia_tri(value)
    if not value:
        return ""

    match = re.match(r"^(?P<number>\d+[^\W_]*)(?P<separator>[-/.])(?P<rest>.+)$", value)
    if not match:
        return value

    rest = re.sub(r"[-/]+", "-", match.group("rest").strip("-/ "))
    return f"{match.group('number')}/{rest}"


def chuan_hoa_so_van_ban_de_so_trung(value) -> str:
    return dinh_dang_so_van_ban_de_ghi(value).lower().replace("/", "-").replace("đ", "d")


def khop_bien_tu_quy_tac(text: str, start: int, end: int) -> bool:
    before = text[start - 1] if start > 0 else ""
    after = text[end] if end < len(text) else ""

    if la_ky_tu_trong_tu(text[start]) and la_ky_tu_trong_tu(before):
        return False
    if la_ky_tu_trong_tu(text[end - 1]) and la_ky_tu_trong_tu(after):
        return False

    return True


def trung_voi_doan_da_chon(start: int, end: int, spans: list[tuple]) -> bool:
    return any(start < old_end and end > old_start for old_start, old_end, _ in spans)


def tim_cac_doan_theo_quy_tac(text: str) -> list[tuple[int, int, str]]:
    spans = []
    for _tieng_viet, tieng_anh, pattern in doc_quy_tac_dich():
        for match in pattern.finditer(text):
            start, end = match.span()
            if not khop_bien_tu_quy_tac(text, start, end):
                continue
            if trung_voi_doan_da_chon(start, end, spans):
                continue
            spans.append((start, end, tieng_anh))

    return sorted(spans, key=lambda item: item[0])


@lru_cache(maxsize=512)
def dich_google_sang_tieng_anh(text: str) -> str:
    return GoogleTranslator(source="vi", target="en").translate(text)


def dich_doan_khong_co_quy_tac(text: str) -> str:
    if not text.strip():
        return text

    prefix = text[: len(text) - len(text.lstrip())]
    suffix = text[len(text.rstrip()) :]
    core = text.strip()
    if not re.search(r"[^\W_]", core, re.UNICODE):
        return text

    try:
        return f"{prefix}{dich_google_sang_tieng_anh(core)}{suffix}"
    except Exception:
        return text


def chuan_hoa_ket_qua_dich(value: str) -> str:
    value = re.sub(r"\s+([,.;:?!])", r"\1", value)
    value = re.sub(r"([(])\s+", r"\1", value)
    value = re.sub(r"\s+([)])", r"\1", value)
    return re.sub(r"\s+", " ", value).strip()


@lru_cache(maxsize=256)
def dich_trich_yeu_sang_tieng_anh(trich_yeu) -> str:
    original = chuan_hoa_gia_tri(trich_yeu)
    if not original:
        return ""

    spans = tim_cac_doan_theo_quy_tac(original)
    if not spans:
        return dich_doan_khong_co_quy_tac(original)

    result = []
    last_end = 0
    for start, end, tieng_anh in spans:
        if start > last_end:
            result.append(dich_doan_khong_co_quy_tac(original[last_end:start]))
        result.append(tieng_anh)
        last_end = end

    if last_end < len(original):
        result.append(dich_doan_khong_co_quy_tac(original[last_end:]))

    return chuan_hoa_ket_qua_dich("".join(result))


def tao_khoa_trung_lap(so_van_ban, ngay_ky, trich_yeu) -> tuple:
    so_van_ban = chuan_hoa_so_van_ban_de_so_trung(so_van_ban)
    ngay_ky = chuan_hoa_gia_tri(ngay_ky).lower()
    trich_yeu = chuan_hoa_gia_tri(trich_yeu).lower()

    if so_van_ban:
        return ("so", so_van_ban, ngay_ky)

    return ("noi_dung", ngay_ky, trich_yeu)


def tach_du_lieu_tu_ten_file(file_name: str) -> tuple[list[str], list[str]]:
    ten_file_khong_duoi, _ = os.path.splitext(file_name)
    ten_file_khong_ngay, ngay_ky = tach_ngay_ky_tu_ten_file(ten_file_khong_duoi)

    match = DOCUMENT_CODE_PATTERN.match(ten_file_khong_ngay)
    if match:
        so_van_ban = dinh_dang_so_van_ban_de_ghi(match.group("code"))
        trich_yeu = chuan_hoa_gia_tri(match.group("content"))
    else:
        so_van_ban = ""
        trich_yeu = ten_file_khong_ngay.strip()

    thieu_du_lieu = []
    if not so_van_ban:
        thieu_du_lieu.append("so_van_ban")
    if not trich_yeu:
        thieu_du_lieu.append("trich_yeu")
    if not ngay_ky:
        thieu_du_lieu.append("ngay_ky")

    return [so_van_ban, ngay_ky, trich_yeu], thieu_du_lieu


def xu_ly_danh_sach_file(duong_dan_folder: str) -> tuple[list[list[str]], list[dict], int]:
    danh_sach_file = [
        f
        for f in os.listdir(duong_dan_folder)
        if os.path.isfile(os.path.join(duong_dan_folder, f))
        and Path(f).suffix.lower() == ".pdf"
        and not f.startswith(".")
    ]

    processed_data = []
    file_loi = []
    for file_name in danh_sach_file:
        data, thieu_du_lieu = tach_du_lieu_tu_ten_file(file_name)
        if thieu_du_lieu:
            file_loi.append({"file_name": file_name, "missing": thieu_du_lieu, "data": data})
            continue

        processed_data.append(data)

    processed_data.sort(key=lambda x: x[0])
    return processed_data, file_loi, len(danh_sach_file)


def dinh_dang_bang(ws) -> None:
    font_family = "Times New Roman"
    header_font = Font(name=font_family, size=11, bold=True, color="000000")
    header_font_red = Font(name=font_family, size=11, bold=True, color="FF0000")
    data_font = Font(name=font_family, size=11, bold=False)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.views.sheetView[0].showGridLines = True
    ws.row_dimensions[1].height = 35

    for col_num, header_text in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.border = cell_border
        cell.alignment = align_center
        cell.font = header_font_red if col_num == TOTAL_COLUMNS else header_font

    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 24
        for col_num in range(1, TOTAL_COLUMNS + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = cell_border
            cell.alignment = align_left if col_num in [1, TRICH_YEU_COL, TRICH_YEU_EN_COL] else align_center

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55


def tao_workbook_moi():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    dinh_dang_bang(ws)
    return wb, ws


def da_co_cot_trich_yeu_tieng_anh(ws) -> bool:
    header = chuan_hoa_header(ws.cell(row=1, column=TRICH_YEU_EN_COL).value)
    return header in {"trích yếu tiếng anh", "trich yeu tieng anh", "trich yeu english", "english summary"}


def dam_bao_cot_trich_yeu_tieng_anh(ws) -> None:
    if not da_co_cot_trich_yeu_tieng_anh(ws):
        ws.insert_cols(TRICH_YEU_EN_COL)

    for col_num, header_text in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_num, value=header_text)

    for row_num in range(2, ws.max_row + 1):
        trich_yeu = ws.cell(row=row_num, column=TRICH_YEU_COL).value
        trich_yeu_en = ws.cell(row=row_num, column=TRICH_YEU_EN_COL).value
        if trich_yeu and not chuan_hoa_gia_tri(trich_yeu_en):
            ws.cell(row=row_num, column=TRICH_YEU_EN_COL, value=dich_trich_yeu_sang_tieng_anh(trich_yeu))


def mo_hoac_tao_workbook(file_excel_output: str):
    excel_path = Path(file_excel_output).expanduser()
    if not excel_path.is_absolute():
        excel_path = Path.cwd() / excel_path

    if not excel_path.exists():
        wb, ws = tao_workbook_moi()
        return wb, ws, excel_path, set()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    dam_bao_cot_trich_yeu_tieng_anh(ws)

    du_lieu_cu = set()
    for row_num in range(2, ws.max_row + 1):
        so_van_ban = ws.cell(row=row_num, column=1).value
        so_van_ban_da_chuan = dinh_dang_so_van_ban_de_ghi(so_van_ban)
        if so_van_ban_da_chuan != chuan_hoa_gia_tri(so_van_ban):
            ws.cell(row=row_num, column=1, value=so_van_ban_da_chuan)

        ngay_ky = ws.cell(row=row_num, column=2).value
        trich_yeu = ws.cell(row=row_num, column=3).value
        du_lieu_cu.add(tao_khoa_trung_lap(so_van_ban, ngay_ky, trich_yeu))

    return wb, ws, excel_path, du_lieu_cu


def append_dong_moi(ws, processed_data: list[list[str]], du_lieu_cu: set) -> tuple[int, int]:
    so_dong_them = 0
    so_dong_trung = 0

    for data in processed_data:
        khoa = tao_khoa_trung_lap(data[0], data[1], data[2])
        if khoa in du_lieu_cu:
            so_dong_trung += 1
            continue

        ws.append([data[0], data[1], data[2], dich_trich_yeu_sang_tieng_anh(data[2]), "", "", ""])
        du_lieu_cu.add(khoa)
        so_dong_them += 1

    return so_dong_them, so_dong_trung


def process_folder_to_excel(
    folder_path: str,
    excel_output: str = "quan_ly_van_ban.xlsx",
    progress_callback: ProgressCallback | None = None,
) -> ProcessingResult:
    folder_path = chuan_hoa_gia_tri(folder_path).strip('"').strip("'")
    excel_output = chuan_hoa_gia_tri(excel_output) or "quan_ly_van_ban.xlsx"

    if not folder_path:
        return ProcessingResult(False, "Chưa chọn thư mục PDF.")

    if not os.path.exists(folder_path):
        return ProcessingResult(False, f"Thư mục không tồn tại: {folder_path}")

    _notify(progress_callback, "Đang quét thư mục PDF...")
    processed_data, file_loi, tong_file = xu_ly_danh_sach_file(folder_path)
    if tong_file == 0:
        return ProcessingResult(True, "Không tìm thấy file PDF nào.", total_files=0)

    if not processed_data:
        return ProcessingResult(
            True,
            "Không có file hợp lệ để ghi vào Excel.",
            total_files=tong_file,
            invalid_files=len(file_loi),
            invalid_items=file_loi,
        )

    _notify(progress_callback, f"Tìm thấy {tong_file} file PDF, {len(processed_data)} file hợp lệ.")

    try:
        wb, ws, excel_path, du_lieu_cu = mo_hoac_tao_workbook(excel_output)
    except PermissionError:
        message = f"Không thể đọc file Excel. Hãy đóng file rồi chạy lại: {excel_output}"
        return ProcessingResult(False, message, total_files=tong_file, valid_files=len(processed_data), invalid_files=len(file_loi), invalid_items=file_loi, system_error=message)
    except Exception as exc:
        message = f"Không thể mở file Excel '{excel_output}': {exc}"
        return ProcessingResult(False, message, total_files=tong_file, valid_files=len(processed_data), invalid_files=len(file_loi), invalid_items=file_loi, system_error=str(exc))

    _notify(progress_callback, "Đang ghi dữ liệu và dịch trích yếu...")
    so_dong_them, so_dong_trung = append_dong_moi(ws, processed_data, du_lieu_cu)
    dinh_dang_bang(ws)

    try:
        wb.save(excel_path)
    except PermissionError:
        message = f"Không thể ghi file Excel. Hãy đóng file rồi chạy lại: {excel_path}"
        return ProcessingResult(False, message, str(excel_path), tong_file, len(processed_data), len(file_loi), so_dong_them, so_dong_trung, file_loi, message)

    message = f"Hoàn tất. Đã lưu Excel tại: {excel_path}"
    return ProcessingResult(
        True,
        message,
        str(excel_path),
        tong_file,
        len(processed_data),
        len(file_loi),
        so_dong_them,
        so_dong_trung,
        file_loi,
    )


def format_result_summary(result: ProcessingResult) -> str:
    lines = [
        result.message,
        f"Đã quét: {result.total_files} file",
        f"Hợp lệ: {result.valid_files} file",
        f"Lỗi dữ liệu: {result.invalid_files} file",
        f"Đã thêm: {result.rows_added} dòng mới",
        f"Trùng lặp: {result.duplicate_rows} dòng",
    ]

    if result.system_error:
        lines.append(f"Lỗi hệ thống: {result.system_error}")

    if result.invalid_items:
        lines.append("File thiếu dữ liệu:")
        for item in result.invalid_items[:10]:
            so_van_ban, ngay_ky, trich_yeu = item["data"]
            lines.append(
                f"- {item['file_name']} | thiếu: {', '.join(item['missing'])} | "
                f"đọc được: so='{so_van_ban}', ngay='{ngay_ky}', trích yếu='{trich_yeu}'"
            )
        if len(result.invalid_items) > 10:
            lines.append(f"... và {len(result.invalid_items) - 10} file khác.")

    return "\n".join(lines)


def run_business_task():
    return "Vui lòng chọn thư mục PDF và file Excel đầu ra."
